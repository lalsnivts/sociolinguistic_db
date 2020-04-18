import datetime
import openpyxl
import os
import re
import sys
from datetime import datetime

from src.preprocessing.ExcelParser import ExcelParser


class ExcelParserFrom2016To2019(ExcelParser):
    worksheet_num = 0

    column_corresp_worksheet = dict()
    column_names_specific_dict = dict()
    objects_parsed_worksheet = dict()

    SPECIFIC_WORKSHEET_NAMES = ["Вопросы",
                                "Владение языками",
                                "Уровень владения",
                                "Периоды жизни"]

    def normalize_text(self, text):
        number = re.compile(r'\d{1,2}\. ')
        text = number.sub('', str(text))
        return text

    def process_column_names_worksheet(self, row, worksheet_index, sheetname):
        unknown_column_names = []
        column_name_worksheet = []
        column_corresp_worksheet = []
        for column_index, cell in enumerate(row):
            column_name = cell.value
            if column_name is None:
                continue
            column_name = self.normalize_text(column_name)
            column_name_worksheet.append(column_name)
            if not self.is_normal_worksheet(sheetname):
                self.column_names_specific_dict[column_index] = column_name
            self.total_column_num += 1
            column_correspondence = self.column_correspondences.get(column_name.lower())
            if column_correspondence is None:
                unknown_column_names.append(column_name)
            else:
                if column_correspondence not in column_corresp_worksheet:
                    column_corresp_worksheet.append(column_correspondence)
        if not self.is_normal_worksheet(sheetname) and worksheet_index != self.worksheet_num:
            column_corresp_worksheet.append('Comments')
        elif not self.is_normal_worksheet(sheetname):
            column_corresp_worksheet.append('Languages')
        self.columns_parsed.append(column_name_worksheet)
        self.column_corresp_worksheet[worksheet_index] = column_corresp_worksheet

    def convert_excel_to_csv(self, filename_in, filename_out=''):
        self.parse_excel(filename_in)
        self.write_to_csv(filename_in)

    def parse_excel(self, filename_in):
        wb = openpyxl.load_workbook(filename=filename_in)
        self.worksheet_num = len(wb.sheetnames) - 1
        for index, sheetname in enumerate(wb.sheetnames):
            self.parse_worksheet(wb[sheetname], sheetname, index)

    def parse_worksheet(self, worksheet, sheetname, worksheet_index):
        self.total_column_num = 0
        self.objects_parsed = []
        sheetname = self.normalize_text(sheetname)
        for row_index, row in enumerate(worksheet.rows):
            if row_index == 0:
                self.process_column_names_worksheet(row, worksheet_index, sheetname)
            elif self.is_normal_worksheet(sheetname) and worksheet_index != self.worksheet_num:
                is_empty_row = self.process_object_data(worksheet_index, row, row_index, False)
                if is_empty_row:
                    break
            elif not self.is_normal_worksheet(sheetname) and worksheet_index != self.worksheet_num:
                self.process_specific_worksheet(worksheet_index, row)
            else:
                self.process_period_worksheet(worksheet_index, row)
        self.objects_parsed_worksheet[worksheet_index] = self.objects_parsed

    def process_specific_worksheet(self, worksheet_index, row):
        fio_custom = ''
        fio_current = ''
        lang_custom = ''
        lang_current = ''
        is_empty_row = True
        object_parsed = dict()
        for column_index, cell in enumerate(row):
            if column_index >= self.total_column_num:
                break
            column_name = self.columns_parsed[worksheet_index][column_index].lower()
            if column_index == 0:
                fio_custom = self.column_correspondences.get(column_name)
                fio_current = self.normalize_value(fio_custom, cell.value)
            elif column_index == 1:
                lang_custom = self.column_correspondences.get(column_name)
                lang_current = self.normalize_value(lang_custom, cell.value)
            elif column_index > 1 and cell.value is not None:
                normalized_name = self.column_correspondences.get(column_name)
                normalized_value = self.normalize_value(normalized_name, cell.value)
                if ',' in normalized_value:
                    normalized_value = normalized_value.replace(',', ';')
                object_parsed[column_index] = dict()
                object_parsed[column_index][fio_custom] = fio_current
                object_parsed[column_index][lang_custom] = lang_current
                object_parsed[column_index][normalized_name] = self.column_names_specific_dict[column_index]
                object_parsed[column_index]['Comments'] = normalized_value
            if cell.value is not None:
                is_empty_row = False
        if not is_empty_row:
            self.objects_parsed.append(object_parsed)

    def process_period_worksheet(self, worksheet_index, row):
        fio_custom = ''
        fio_current = ''
        is_empty_row = True
        object_parsed = dict()
        for column_index, cell in enumerate(row):
            if column_index >= self.total_column_num:
                break
            column_name = self.columns_parsed[worksheet_index][column_index]
            if column_index == 0:
                fio_custom = self.column_correspondences.get(column_name)
                fio_current = self.normalize_value(fio_custom, cell.value)
            elif column_index > 0 and cell.value is not None:
                normalized_name = self.column_correspondences.get(column_name)
                normalized_value = self.normalize_value(normalized_name, cell.value)
                if ',' in normalized_value:
                    normalized_value = normalized_value.replace(',', ';')
                object_parsed[column_index] = dict()
                object_parsed[column_index][fio_custom] = fio_current
                object_parsed[column_index][normalized_name] = self.column_names_specific_dict[column_index]
                object_parsed[column_index]['Languages'] = normalized_value
            if cell.value is not None:
                is_empty_row = False
        if not is_empty_row:
            self.objects_parsed.append(object_parsed)

    def write_to_csv(self, filename_in):
        wb = openpyxl.load_workbook(filename=filename_in)
        for worksheet_index, sheetname in enumerate(wb.sheetnames):
            sheetname = self.normalize_text(sheetname)
            filename_out = self.prepare_filename_out(filename_in, sheetname)
            print(filename_out)
            with open(filename_out, 'w', encoding='utf-8', newline='') as fout:
                fout.write(self.create_csv_headers_idx(worksheet_index) + self.LINE_SEPARATOR)
                for object_parsed in self.objects_parsed_worksheet[worksheet_index]:
                    if self.is_normal_worksheet(sheetname):
                        fout.write(self.create_csv_line_idx(object_parsed, worksheet_index) + self.LINE_SEPARATOR)
                    else:
                        for value in object_parsed.values():
                            fout.write(self.create_csv_line_idx(value, worksheet_index) + self.LINE_SEPARATOR)

    def prepare_filename_out(self, filename_in, sheetname):
        cur_dir = os.getcwd()
        folder = os.path.splitext(os.path.split(filename_in)[1])[0]
        final_dir = os.path.join(cur_dir, folder)
        if not os.path.exists(final_dir):
            os.makedirs(final_dir)
        filename_out = r'%s\%s_%s.csv' % (final_dir, folder, sheetname)
        return filename_out

    def create_csv_headers_idx(self, worksheet_index):
        return self.COLUMN_SEPARATOR.join(self.column_corresp_worksheet[worksheet_index])

    def create_csv_line_idx(self, object_parsed, worksheet_index):
        line_csv = ''
        for column_name in self.column_corresp_worksheet[worksheet_index]:
            line_csv += str(object_parsed.get(column_name, '')) + self.COLUMN_SEPARATOR
        return line_csv.strip()


def main():
    excel_parser = ExcelParserFrom2016To2019(sys.argv[1])
    for arg in sys.argv[2:]:
        excel_parser.convert_excel_to_csv(arg)


if __name__ == '__main__':
    main()
